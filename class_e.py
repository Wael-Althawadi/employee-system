from tkinter import *
from PIL import Image,ImageTk
from tkinter import ttk
import pandas as pd
import numpy
from tkinter import filedialog
from openpyxl import workbook,load_workbook
from tkinter import messagebox
from tkinter.font import Font

class Sabratha_u:
    def __init__(self,master,title,geometry):
        global i_b
        global i_b2
        global i_b3
        global i_b4
        
        self.my_font = Font(
        family = 'Times',
        size = 30,
        weight = 'bold',
        slant = 'roman',
        underline = 0,
        overstrike = 0
        )
        
        master.geometry(geometry)
        master.title(title)
        master.configure(bg = "white")
        self.B_frame = Frame(master,bg = "#80b5b0")
        self.B_frame.place(height = 70,width = 450,x = 430,y = 0)
        
        self.l = Label(self.B_frame,text = "Sabratha University",fg = "white",font = self.my_font,bg = "#80b5b0")
        self.l.place(width = 450, height =40, x= 0, y = 10)
        
        self.B_frame2 = Frame(master,bg = "white")
        self.B_frame2.place(height = 570,width = 1295,x = 0,y = 70)
        
        
        add_img = Image.open("/home/wael/un-sys/add-icon.jpg")
        i_b = ImageTk.PhotoImage(add_img)
        self.b_i = Button(master,image = i_b,bg = "white",command = self.add_user )
        self.b_i.place(x = 1000,y = 120)
        
        
        del_img = Image.open("/home/wael/un-sys/delete-icon.jpg")
        i_b2 = ImageTk.PhotoImage(del_img)
        self.b_i2 = Button(master,image = i_b2,bg = "white",command = self.delete_user )
        self.b_i2.place(x = 750,y = 120)

        
        srch_img = Image.open("/home/wael/un-sys/search-icon.jpg")
        i_b3 = ImageTk.PhotoImage(srch_img)
        self.b_i3 = Button(master,image = i_b3,bg = "white",command = self.srch_user)
        self.b_i3.place(x = 400,y = 120)

        
        upd_img = Image.open("/home/wael/un-sys/update-icon.jpg")
        i_b4 = ImageTk.PhotoImage(upd_img)
        self.b_i4 = Button(master,image = i_b4,bg = "white")
        self.b_i4.place(x = 150,y = 120)
        
        self.menu2 = Menu(master,bg = "#80b5b0")
        self.sub_m = Menu(self.menu2,tearoff = 1) # or you can set tearoff to 0 to remove the line
        self.sub_m.add_command(label = "add employees",command = self.add_user)
        self.sub_m.add_command(label = "delete employees",command = self.delete_user)
        self.sub_m.add_command(label = "employees search",command = self.srch_user)
        self.sub_m.add_command(label = "update employees")
        self.sub_m.add_separator()
        
        self.menu2.add_cascade(label = "edit",menu = self.sub_m)
        self.menu2.add_command(label = "exit",command = master.destroy)
        master.config(menu = self.menu2) 
        
        self.bottom_f = Frame(master,bg = "#80b5b0",bd = 0,relief = RIDGE)
        self.bottom_f.place(width = 1300,height = 25 ,x = 0,y = 653)
        
  
        
    def add_user(self):
        global pic
        global my_tree
        
        my_tree = ""
        
        root_add = Toplevel()
        root_add.geometry(f"{root_add.winfo_screenwidth()}x{root_add.winfo_screenheight()}")
        root_add.configure(bg = "#aab7b8")
        
        add_usr_ = Frame(root_add,bg = "#aab7b8",padx = 400)
        add_usr_.pack()

        add_pic = Image.open("/home/wael/un-sys/employee-icon.jpg")
        pic = ImageTk.PhotoImage(add_pic)

        l_img = Label(root_add,image = pic,bg = "#aab7b8",bd = 0)
        l_img.place(x = 1000,y = 30 )
        
        l_img2 = Label(root_add,image = pic,bg = "#aab7b8",bd = 0)
        l_img2.place(x = 100,y = 30 )
        
        l_1 = Label(add_usr_,text = "Full Name",padx = 30,bg = "#F5F5F5")
        l_1.grid(row = 0 ,column = 0)

        en_1 = Entry(add_usr_,borderwidth = 3)
        en_1.grid(row = 0,column = 1,ipadx = 90,ipady = 5)

        l_2 = Label(add_usr_,text = "age",padx = 30,bg = "#F5F5F5")
        l_2.grid(row = 1 ,column = 0)

        en_2 = Entry(add_usr_,borderwidth = 3)
        en_2.grid(row = 1,column = 1,ipadx = 90,ipady = 5)

        l_3 = Label(add_usr_,text = "ID",padx = 30,bg = "#F5F5F5")
        l_3.grid(row = 2 ,column = 0)

        en_3 = Entry(add_usr_,borderwidth = 3)
        en_3.grid(row = 2,column = 1,ipadx = 90,ipady = 5)

        l_4 = Label(add_usr_,text = "password",padx = 30,bg = "#F5F5F5")
        l_4.grid(row = 3 ,column = 0)

        en_4 = Entry(add_usr_,borderwidth = 3)
        en_4.grid(row = 3,column = 1,ipadx = 90,ipady = 5)

        s__v = StringVar()
        Radiobutton(add_usr_,text = "male",variable = s__v,value = 1).place(x = 130,y = 128)
        Radiobutton(add_usr_,text = "female",variable = s__v,value = 2).place(x = 300,y = 128)

        

        def go_add():
            
            if my_tree != "" :
                if en_1.get() and en_2.get() != "":
                    my_tree.insert("","end",values = (en_1.get(),en_2.get()))

                    wb = load_workbook(file_n)
                    ws = wb.active
                    ws.append([en_1.get(),en_2.get()])
                    wb.save(file_n)

                    messagebox.showinfo("message","succesfully added")
                    en_1.delete(0,END)
                    en_2.delete(0,END)
                else:
                    messagebox.showerror("error","check fields")
            else:
                messagebox.showerror("error","open file firist")

        add_go = Button(add_usr_,text = "ADD USER",bg = "#3498db",fg = "white",padx = 138,pady = 10,command = go_add)
        add_go.grid(row = 4 ,column = 1,pady = 25,rowspan = 4)


        def opn_file():
            global my_tree
            global df_rows_
            global file_n

            file_types = (("excel files","*.xlsx"),("all files","*.*"))
            file_n = filedialog.askopenfilename(initialdir = "/home/",title = "title",filetypes = file_types)

            if file_n:

                df = pd.read_excel(file_n)

                frame_add = Frame(root_add,bg = "#aab7b8")
                frame_add.place(width = 1291,x = 1,y = 250)

                my_tree = ttk.Treeview(frame_add)

                style_ = ttk.Style(frame_add)
                style_.theme_use('clam')

                my_tree["column"] = list(df.columns)
                my_tree["show"] = "headings"

                for col_ in df.columns:
                    my_tree.heading(col_,text=col_)

                df_rows_ = df.to_numpy().tolist()
                for row_ in df_rows_:
                    my_tree.insert("", "end",values= row_)


                my_tree.pack()
            else:
                messagebox.showwarning("warning","you did not choose a csv file")

        menu2 = Menu(root_add)
        menu2.add_command(label = "open file",command = opn_file)
        root_add.config(menu = menu2)



        
    def delete_user(self):
        global pic2
        del_usr = Toplevel()
        del_usr.geometry(f"{del_usr.winfo_screenwidth()}x{del_usr.winfo_screenheight()}")
        del_usr.configure(bg = "#aab7b8")
        
        del_usr_ = Frame(del_usr,bg = "#aab7b8",padx = 400)
        del_usr_.pack()
    
        add_pic2 = Image.open("/home/wael/un-sys/employee-icon.jpg")
        pic2 = ImageTk.PhotoImage(add_pic2)

        
        l_img_ = Label(del_usr,image = pic2,bg = "#aab7b8",bd = 0)
        l_img_.place(x = 1000,y = 30 )
        
        l_img_2 = Label(del_usr,image = pic2,bg = "#aab7b8",bd = 0)
        l_img_2.place(x = 100,y = 30 )
        
        l__1 = Label(del_usr_,text = "Full Name",padx = 30,bg = "#F5F5F5")
        l__1.grid(row = 0 ,column = 0)

        en__1 = Entry(del_usr_,borderwidth = 3)
        en__1.grid(row = 0,column = 1,ipadx = 90,ipady = 5)

        l__2 = Label(del_usr_,text = "age",padx = 30,bg = "#F5F5F5")
        l__2.grid(row = 1 ,column = 0)

        en__2 = Entry(del_usr_,borderwidth = 3)
        en__2.grid(row = 1,column = 1,ipadx = 90,ipady = 5)

        l__3 = Label(del_usr_,text = "ID",padx = 30,bg = "#F5F5F5")
        l__3.grid(row = 2 ,column = 0)

        en__3 = Entry(del_usr_,borderwidth = 3)
        en__3.grid(row = 2,column = 1,ipadx = 90,ipady = 5)

        l__4 = Label(del_usr_,text = "password",padx = 30,bg = "#F5F5F5")
        l__4.grid(row = 3 ,column = 0)

        en__4 = Entry(del_usr_,borderwidth = 3)
        en__4.grid(row = 3,column = 1,ipadx = 90,ipady = 5)

        s__v2 = StringVar()
        Radiobutton(del_usr_,text = "male",variable = s__v2,value = 1).place(x = 130,y = 128)
        Radiobutton(del_usr_,text = "female",variable = s__v2,value = 2).place(x = 300,y = 128)
        
        
        def del_go():
            
            if my_tree2 != "" :
                wb = load_workbook(file_n)
                ws = wb.active
                for row in ws.iter_rows():
                    for cell in row:
                        if en__1.get() in str(cell.value):
                            ws.delete_rows(cell.row)
                wb.save(file_n)
                    
                en__1.delete(0,END)
                en__2.delete(0,END)
                x = my_tree2.selection()
                
                if x :
                    for i in x:
                        my_tree2.delete(x)
                        messagebox.showinfo("message","succefuly deleted")
                else:
                    messagebox.showerror("error","select user to delete")
            else:
                messagebox.showerror("error","open file firist")
                
        add_go = Button(del_usr_,text = "DELETE USER",bg = "#3498db",fg = "white",padx = 130,pady = 10,command = del_go)
        add_go.grid(row = 4 ,column = 1,pady = 5)
        
        def srch_go():
            
            if my_tree2 != "":
                if en__1.get() == "" :
                    x2 = my_tree2.selection()
                    items = my_tree2.item(x2)
                                                                ########################################
                    en__1.insert(0,str(items["values"][0]))
                    en__2.insert(0,int(items["values"][1]))
                elif en__1.get() != "":
                    for child in my_tree2.get_children():
                        if en__1.get() in my_tree2.item(child)["values"]:
                            my_tree2.selection_set(child)
                                
                            en__1.delete(0,END)
                            en__1.insert(0,str(child["values"][0]))
                            en__2.insert(0,int(child["values"][1]))
            
                
        
                
        add_go = Button(del_usr_,text = "Get info",bg = "#3498db",fg = "white",padx = 130,pady = 10,command = srch_go)
        add_go.grid(row = 5 ,column = 1,pady = 5)
        
        
        def opn_file2():
            global my_tree2
            global df_rows_2
            global file_n

            file_types = (("excel files","*.xlsx"),("all files","*.*"))
            file_n = filedialog.askopenfilename(initialdir = "/home/",title = "title",filetypes = file_types)

            if file_n:

                df2 = pd.read_excel(file_n)

                frame_add2 = Frame(del_usr,bg = "#aab7b8")
                frame_add2.place(width = 1291,x = 1,y = 250)

                my_tree2 = ttk.Treeview(frame_add2)

                style_2 = ttk.Style(frame_add2)
                style_2.theme_use('clam')

                my_tree2["column"] = list(df2.columns)
                my_tree2["show"] = "headings"

                for col_ in df2.columns:
                    my_tree2.heading(col_,text=col_)

                df_rows_2 = df2.to_numpy().tolist()
                for row_ in df_rows_2:
                    my_tree2.insert("", "end",values= row_)


                my_tree2.pack()
            else:
                messagebox.showwarning("warning","you did not choose a csv file")

        menu3 = Menu(del_usr)
        menu3.add_command(label = "open file",command = opn_file2)
        del_usr.config(menu = menu3)

        
        
    def srch_user(self):
        global pic3
        global my_tree3
                
        root_srch = Toplevel()
        root_srch.geometry(f"{root_srch.winfo_screenwidth()}x{root_srch.winfo_screenheight()}")
        root_srch.configure(bg = "#aab7b8")
        
        srch_usr = Frame(root_srch,bg = "#aab7b8",padx = 400,bd = 1,relief = "raised")
        srch_usr.pack()

        add_pic3 = Image.open("/home/wael/un-sys/employee-icon.jpg")
        pic3 = ImageTk.PhotoImage(add_pic3)

        l_img3 = Label(root_srch,image = pic3,bg = "#aab7b8",bd = 0)
        l_img3.place(x = 1000,y = 25 )
        
        l_img_3 = Label(root_srch,image = pic3,bg = "#aab7b8",bd = 0)
        l_img_3.place(x = 100,y = 25 )
        
        l_1_ = Label(srch_usr,text = "Full Name",padx = 30,bg = "#F5F5F5")
        l_1_.grid(row = 0 ,column = 0)

        en_1_ = Entry(srch_usr,borderwidth = 3)
        en_1_.grid(row = 0,column = 1,ipadx = 90,ipady = 5)

        l_2_ = Label(srch_usr,text = "age",padx = 30,bg = "#F5F5F5")
        l_2_.grid(row = 1 ,column = 0)

        en_2_ = Entry(srch_usr,borderwidth = 3)
        en_2_.grid(row = 1,column = 1,ipadx = 90,ipady = 5)

     
        
        
        def search():
            if my_tree3 and en_1_.get() != "" :
                for child in my_tree3.get_children():
                    if en_1_.get() in my_tree3.item(child)["values"]:
                        my_tree3.selection_set(child)
                        
                        en_1_.insert(0,str(child["values"][0]))
                        en_2_.insert(0,int(child["values"][1]))
                    else:
                        messagebox.showerror("error","user not found!")
            else:
                messagebox.showerror("error","check fields ,or file not open")
        
        search_go = Button(srch_usr,text = "USER SEARCH",bg = "#3498db",fg = "white",padx = 130,pady = 10,command = search)
        search_go.grid(row = 4 ,column = 1,pady = 5)
        
        
        def opn_file3():
            global my_tree3
            global df_rows_3
            global file_n

            file_types = (("excel files","*.xlsx"),("all files","*.*"))
            file_n = filedialog.askopenfilename(initialdir = "/home/",title = "title",filetypes = file_types)

            if file_n:

                df3 = pd.read_excel(file_n)

                frame_add3 = Frame(root_srch,bg = "#aab7b8")
                frame_add3.place(width = 1291,x = 1,y = 250)

                my_tree3 = ttk.Treeview(frame_add3)

                style_3 = ttk.Style(frame_add3)
                style_3.theme_use('clam')

                my_tree3["column"] = list(df3.columns)
                my_tree3["show"] = "headings"

                for col_ in df3.columns:
                    my_tree3.heading(col_,text=col_)

                df_rows_3 = df3.to_numpy().tolist()
                for row_ in df_rows_3:
                    my_tree3.insert("", "end",values= row_)


                my_tree3.pack()
            else:
                messagebox.showwarning("warning","you did not choose a csv file")
                
        menu4 = Menu(root_srch)
        menu4.add_command(label = "open file",command = opn_file3)
        root_srch.config(menu = menu4)
